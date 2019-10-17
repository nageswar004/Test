from geopy.geocoders import Nominatim
from openpyxl import load_workbook

from django.shortcuts import render
from django.http import HttpResponseRedirect
from django.core.files.storage import FileSystemStorage

from App.models import Document
from Test.settings import BASE_DIR


def home(request):
    documents = Document.objects.all()
    return render(request, 'home.html', {'documents': documents})


def simple_upload(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)

        if 'xlsx' not in filename:
            return HttpResponseRedirect("/uploads/simple/")

        uploaded_file_url = fs.url(filename)
        file_path = BASE_DIR + uploaded_file_url
        wb = load_workbook(file_path)
        sheet = wb.active
        max_row, max_col = sheet.max_row, sheet.max_column
        geolocator = Nominatim()
        for row in range(1, max_row + 1):
            try:
                loc = geolocator.geocode(sheet.cell(row, 2).value + ',' + sheet.cell(row, 3).value)
                sheet.cell(row, max_col + 1, loc.latitude)
                sheet.cell(row, max_col + 2, loc.longitude)
            except Exception as e:
                sheet.cell(row, max_col + 1, 'latitude')
                sheet.cell(row, max_col + 2, 'longitude')
                continue
        wb.save(file_path)

        return render(request, 'simple_upload.html', {
            'uploaded_file_url': uploaded_file_url
        })
    return render(request, 'simple_upload.html')
